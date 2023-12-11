from ctypes import alignment
from json import load
from cv2 import BORDER_DEFAULT
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill,Alignment,Border, Side

Workbook = None
dest_filename =""
sheet = None

#Workbook.save(dest_filename)

all_mismatch_testcount={}
total_mismatch_test_count = 0  
redFill = PatternFill(start_color='F2DCDB',
                   end_color='F2DCDB',
                   fill_type='solid')
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))


def get_mismatch_count_for_eachmoi(workbook_obj):   
        try: 
            #To find mismatched test testcount for each MOI and create dictionary
            #Ignore summary sheet
            all_sheet_names = workbook_obj.sheetnames   
            all_sheet_names.remove("Summary")
            for sheet_name in all_sheet_names :
                sheet_cells = []
                sh =workbook_obj[sheet_name]

            #To read mismatched count for each row 
                for rows in sh.iter_rows():
                    row_cells = []
                    for cell in rows:
                        row_cells.append(cell.value)
                    # create list which will contain all the row data's as tuple format
                    sheet_cells.append(tuple(row_cells))

                #To find row data which contains value as "FALSE"     
                mismatch_count = 0    
                for indidual_row_data in sheet_cells :               
                    if "FALSE" in indidual_row_data:
                        mismatch_count = mismatch_count + 1

                #create dictionary which contains each MOI and its mismatch count        
                all_mismatch_testcount[sheet_name] = mismatch_count              

        except Exception as e:
            print(str(e))   

def get_total_mismatch_count():
    try:
        #To count total value of all mismatched data 
        count_val = 0
        for key, value in all_mismatch_testcount.items():
            count_val = count_val+value
        return count_val 

    except Exception as e:
        print(str(e)) 

def update_header(sheet,ws,cnt):
    #inser a row in header for the total mismatch count
    sheet.insert_rows(16)
    sheet['A16'] =  "Total Mismatch"
    sheet['B16'] = cnt
    sheet['B16'].alignment = Alignment(horizontal='left')
    sheet.merge_cells(start_row=16, start_column=2, end_row=16, end_column=11)
    sheet['A16'].fill = redFill
    sheet['A16'].border = thin_border
    sheet['B16'].fill = redFill
    sheet['B16'].border = thin_border
def update_detail(sheet,ws):
    sheet['C17']="Mismatch"
    i=1
    for col in sheet['B']:
        if col.value in all_mismatch_testcount:
            sheet['C'+str(i)] = all_mismatch_testcount[col.value]
        i+=1
    
def add_mismatch(path):
    Workbook = load_workbook(filename=path)
    dest_filename = path
    sheet = Workbook["Summary"]

    get_mismatch_count_for_eachmoi(Workbook)
    total_mismatch_test_count=get_total_mismatch_count()
    update_header(sheet,Workbook,total_mismatch_test_count)
    update_detail(sheet,Workbook)
    Workbook.save(dest_filename)