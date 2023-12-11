import os, sys
import json
import io
import enum
import operator
import datetime,shutil 
import glob
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
import xlsxwriter
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from operator import is_not
from functools import partial

report_result_path="C:\\GRL\\C2_Comparison_Tool_Report"

class Update_Summary_Xls_Report:

    def __init__(self):
       
        self.conclusion_sum=  None
        self.project_name = None
        self.latest_summary_xls_file = None
                    
    def get_latest_summary_xls_file(self):
        self.latest_report = max(glob.glob(os.path.join(report_result_path, '*')), key=os.path.getmtime)  
        # Get the list of all files in directory tree at given path
        self.list_of_files = list()
        for (dirpath, dirnames, filenames) in os.walk(self.latest_report):
            self.list_of_files += [os.path.join(dirpath, file) for file in filenames]     
        #get xls summary output file 
        for xls_file in self.list_of_files:
            if '.xlsx' in xls_file:    
                #xls file should contain user given project name along with (_with_Ref) name extension             
                if '_With_Ref' in xls_file and self.project_name in xls_file:
                    #avoid xls file with '.~lock.' extension
                    if '.~lock.' not in xls_file:
                        self.latest_summary_xls_file = xls_file
                        print("output xls filename based on user config port::  "+xls_file)

  
    def write_conclusion_summary_data(self,conclusion_sum,proj_name):
        #assign instance variable to conclusion sum
        self.conclusion_sum= conclusion_sum
        #assign project name
        self.project_name = proj_name
        #To call all the method and update xls file
        self.get_latest_summary_xls_file()
        self.modify_xls_file()
      

    def modify_xls_file(self):
        try:
            bg_clr_fill = PatternFill(start_color='F0E6E5',end_color='FFFFFF',fill_type='solid') 
            thin = Side(border_style="thin", color="303030")
            black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
            #To create object for xls file
            wb = load_workbook(filename = self.latest_summary_xls_file)
            #To create object for first sheet(Summary Sheet)
            ws = wb.worksheets[0]
            #get last row index in the xls file
            max_row_data = int()
            max_row_data = ws.max_row+1   
            #To add summary conclusion in first col
            ws.cell(max_row_data,1).value = "Selected Moi and Testcase Status"
            ws.cell(max_row_data,1).fill =  bg_clr_fill
            ws.cell(max_row_data,1).border = black_border     
            #To add moi and testcase summary conclusion in second col
            ws.cell(max_row_data,2).value = self.conclusion_sum
            ws.cell(max_row_data,2).fill =  bg_clr_fill
            ws.cell(max_row_data,2).border = black_border
            #To merge second col data into particular range(example :B15:K15)
            merge_start_range = '{}{}'.format('B',max_row_data)
            merge_end_range = '{}{}'.format('K',max_row_data)
            merge_total_range ='{}:{}'.format(merge_start_range,merge_end_range)
            ws.merge_cells(merge_total_range) 
            #save xls file
            wb.save(self.latest_summary_xls_file)             
            #ws.cell(row_data,2).alignment = Alignment(wrapText=True)
            
        except Exception as e:
            print("Input  Xls File is Not Found" + str(e))
     


def main(): 
    pass

if __name__ == "__main__":
    main()




                 
               
   
        
    
