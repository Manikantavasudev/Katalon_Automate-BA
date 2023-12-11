
import os, sys,glob, datetime
import json
import io
import enum
import time
import operator
import shutil
import openpyxl
from openpyxl import load_workbook
import xlsxwriter
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from operator import is_not
from functools import partial

report_path = "C:\\GRL\\USBPD-C2-Browser-App\\Report\\TempReport"
report_result_path="C:\\GRL\\C2_Comparison_Tool_Report"

class CreateXlsSummaryReport:

    def __init__(self,port_name):

        #input port name
        self.port_name = port_name

        #get latest input xls file based on selected port
        self.latest_report = None
        self.latest_port_xls_file = None
        self.get_latest_port_xls_file(self.port_name)

        #To create input xls files common  object 
        self.workbook_obj = load_workbook(self.latest_port_xls_file)

        #To create Output xls files common object
        self.temp_file_path = self.latest_report+"\\Temp"+"\\Summary_Report.xlsx"
        self.temp_up_path =self.latest_report+"\\Temp_Report1.xlsx"
        self.summary_file_path = self.latest_report+"\\Summary-"+self.port_name+".xlsx"
        self.op_workbook_obj = xlsxwriter.Workbook(self.temp_file_path)
        self.op_worksheet_obj = self.op_workbook_obj.add_worksheet("Summary")

        

        # To set xls background format
        self.head_format = None
        self.odd_format = None
        self.even_format = None
        
        #To define common data variable
        self.ignore_sheetname = "Summary"
        self.all_sheet_names = None
        self.all_mismatch_testcount = {}
        self.total_mismatch_test_count = None       
        self.summary_sheet_data = None
        self.modify_summary_list = []
        self.updated_summary_with_folderPath = []

        self.up_summary_sheet_data = []
        self.fil_sum_data=[]

        #set Test Execution Info deatils 
        self.set_test_info_order = []
        self.test_execution_info_list = []
        
        # To define dict values for modify summary deatils
        self.update_sum_info_rowwise = {"C2 Sno":"C2 Serial Number","Five Port":"True","Test Er":"Test Engineer",
                                        "Port Type":"DUT Type","SW Ver":"Software Version","FW Ver":"Firmware Version",
                                        "Folder Abs Path":"Folder Abs Path","Folder Name":"Result Folder Name"} 
        self.ignore_sum_info = ["Dev Type"]

        #Set TestSetup Information  
        self.test_setup_info_order = {"Test Setup Information":0,"Date":1,"C2 Serial Number":2,
                                    "Five Port":3,"Test Engineer":4,"Software Version":5,"Firmware Version":6,
                                    "PD Spec Rev":7,"Result Folder Name":8,"Folder Abs Path":9}

        self.header_format = ['Folder Name', 'MOI Name', 'Total', 'Mismatch', 
                            'Pass', 'Fail', 'Incomp', 'Not Exe', 'Not App', 'Others', 'Not Sel']
       
        #To call all the methods
        self.tot_count = 0
        self.write_data()

 
    """ To get latest Comparison five port report folder based on latest timestamp 
    and get latest xls summary file,
    xls files will  present based on port name in following path 'C:\\GRL\\C2_Comparison_Tool_Report\\FivePort* etc..' """    

    def get_latest_port_xls_file(self,port_name):
        try:
            self.latest_report = max(glob.glob(os.path.join(report_result_path, '*')), key=os.path.getmtime)   
            # Get the list of all files in directory tree at given path
            list_of_files = list()
            for (dirpath, dirnames, filenames) in os.walk(self.latest_report):
                list_of_files += [os.path.join(dirpath, file) for file in filenames]   
            #get xls output file based on selected port
            for xls_file in list_of_files:
                if '.xlsx' in xls_file:                  
                    if '_With_Ref' in xls_file and port_name in xls_file:
                        #avoid xls file with '.~lock.' extension
                        if '.~lock.' not in xls_file:
                            self.latest_port_xls_file = xls_file
                            print("output xls filename based on user config port::  "+xls_file) 
                            #shutil.copy2(xls_file, "dest-demo.xlsx")   
        except Exception as e:
            print(str(e))                      

    def write_data(self): 
        try:   
            self.get_all_sheet_names(self.workbook_obj)
            self.get_mismatch_count_for_eachmoi(self.workbook_obj)
            self.get_total_mismatch_count()
            self.get_summary_details(self.workbook_obj)
            self.modify_summary_details()
            self.modify_summary_folder_path()
            self.set_order_for_summary_details()
            self.select_format()      
            self.xls_testInformation_writter() 
        except Exception as e:
            print(str(e))     
        
       
    def get_all_sheet_names(self,workbook_obj): 
        try:
            #To read all the sheet names as a list from input xls files     
            self.all_sheet_names = workbook_obj.sheetnames
        except Exception as e:
            print(str(e)) 


    def get_mismatch_count_for_eachmoi(self,workbook_obj):   
        try: 
            #To find mismatched test testcount for each MOI and create dictionary
            #Ignore summary sheet
            self.all_sheet_names.remove(self.ignore_sheetname)    
            for sheet_name in self.all_sheet_names :
                sheet_cells = []
                sh =workbook_obj[sheet_name]

                #To read mismatched count for each row 
                for rows in sh.iter_rows():
                    row_cells = []
                    for cell in rows:
                        row_cells.append(cell.value)
                    # create list which will contain all the row data's as tuple format
                    sheet_cells.append(tuple(row_cells))

                #To find row data which contains value as "FALSE"     
                mismatch_count = 0    
                for indidual_row_data in sheet_cells :               
                    if "FALSE" in indidual_row_data:
                        mismatch_count = mismatch_count + 1

                #create dictionary which contains each MOI and its mismatch count        
                self.all_mismatch_testcount[sheet_name] = mismatch_count             
                #print(self.all_mismatch_testcount)  
        except Exception as e:
            print(str(e))         


    def get_total_mismatch_count(self):
        try:
            #To count total value of all mismatched data 
            count_val = 0
            for key, value in self.all_mismatch_testcount.items():
                count_val = count_val+value
            self.total_mismatch_test_count = count_val  
        except Exception as e:
            print(str(e)) 


    def get_summary_details(self,workbook_obj):      
        #To get summary details data
        wk =workbook_obj["Summary"]   

        sum_sheet_cells = []    
        #iterate each row and convert the each row data's as a tuple format  
        for rows in wk.iter_rows():
            row_cells = []
            for cell in rows:
                row_cells.append(cell.value)
                # create list which will contain all the row data's as tuple format
            sum_sheet_cells.append(tuple(row_cells))   

        self.summary_sheet_data = sum_sheet_cells
        #print(self.summary_sheet_data)        


    def modify_summary_details(self):
        #read all the summary details 
        for row_data in self.summary_sheet_data :
            #convert tuple as list and modify the list data's
            self.alter_summary_data  = list(row_data)  
            #Remove Run values from list
            if "Run" or "RUN_" in  self.alter_summary_data:
                self.alter_summary_data.pop(2)

            #Add Mismatch column in list ,mismatch column will present next to the Total column
            if "Total" in self.alter_summary_data :
                self.alter_summary_data.insert(3,"Mismatch")

            #Add Mismatch count for Each MOI 
            for sheet_name in self.all_sheet_names:
                if sheet_name in self.alter_summary_data:
                    moi_count  = self.all_mismatch_testcount[sheet_name]
                    self.alter_summary_data.insert(3,moi_count)     

            #To Remove unused repeated row values(Dev Type) 
            if  self.ignore_sum_info[0] in self.alter_summary_data:
                self.alter_summary_data.clear()

            """Modify following summary Information 
            Five Port,C2 Sno,Test Er,Port Type,SW Ver,FW Ver,Folder Name  and remove dev type """
            for sheet_key in self.alter_summary_data:
                try:
                    for sum_key,sum_value in self.update_sum_info_rowwise.items():
                        if sum_key in sheet_key:
                            #To modify fiveport key value as True
                            if sum_key == "Five Port":
                                self.alter_summary_data[1]="True"                               
                            else:  
                                #To ignore Folder name  key value change
                                if "MOI Name" in self.alter_summary_data:
                                    pass
                                #To modify C2 Sno,Test Er,Port Type,SW Ver,FW Ver,Folder Name values
                                else:    
                                    self.alter_summary_data[0]=sum_value
                except Exception as e:
                    pass     
            #Create list which contains all the modified summary data   
            self.modify_summary_list.append(self.alter_summary_data)                
    

    def modify_summary_folder_path(self):
        #To split result folder abs path
        head_tail = os.path.split(self.latest_report)
        try:
            for sum_list in self.modify_summary_list:
                #To update Result Folder 
                if "Result Folder Name" in sum_list:
                    sum_list[1] =  head_tail[1]

                #To update Folder Abs path    
                elif "Folder Abs Path" in sum_list:   
                    sum_list[1] = self.latest_report
                else:
                    pass    
                self.updated_summary_with_folderPath.append(sum_list)
            #print(self.updated_summary_with_folderPath) 
        except Exception as e:
            pass       


    def set_order_for_summary_details(self) :
        #Set TestSetup Information
        for each_summary_list in self.updated_summary_with_folderPath:
            for key, val in self.test_setup_info_order.items():                   
                if key in each_summary_list:                 
                    self.set_test_info_order.append(each_summary_list)                     
        print(self.set_test_info_order)      

        #Set TestExecution Information 
        for summary_list in self.modify_summary_list:
            if summary_list not in self.set_test_info_order:
                if len(summary_list) > 0: 
                    self.test_execution_info_list.append(summary_list)    

        #To add summary details and xls report 
        add_sum_list =["Summary:",self.latest_port_xls_file] 
        #self.test_execution_info_list.append(add_sum_list)           
        #print(self.test_execution_info_list)    


    def select_format(self): 
        self.head_format = self.op_workbook_obj.add_format(
            {'bg_color': '#702D2F', 'border': 1, 'font_color': '#FFFFFF', 'align': 'center',
                 'valign': 'vcenter', 'border_color': '#FFFFFF'})
        self.odd_format = self.op_workbook_obj.add_format(
            {'bg_color': '#F0E6E5', 'border': 1})      
        self.even_format = self.op_workbook_obj.add_format(
            {'bg_color': '#702D2F', 'border': 1,'font_color': '#FFFFFF'})   


    def xls_testInformation_writter(self):         
        if os.path.isfile(self.temp_file_path): 
            self.get_updated_sum_details()
            self.over_write_xls()
            shutil.copy2(self.temp_file_path, self.summary_file_path)
            
        else:              
            self.op_worksheet_obj.merge_range(0, 0, 0, 10, "Test Setup Information",self.head_format)
            row = 1
            col = 0                                
            for test_info_list in self.set_test_info_order:
                for each_val in test_info_list:    
                    if col == 1:           
                        self.op_worksheet_obj.merge_range(row, col,row,col+9, each_val, self.odd_format)   
                    else:
                        self.op_worksheet_obj.write(row, col, each_val, self.odd_format)  
                        self.op_worksheet_obj.set_column('A:A', 30)
                    col += 1 
                row += 1
                col =0
            print(row)   
            self.xls_testExeution_info_writter(row)   
            self.op_workbook_obj.close() 
            shutil.copy2(self.temp_file_path, self.summary_file_path)        


    def xls_testExeution_info_writter(self,tot_row_count):   
        row = tot_row_count+1
        col = 0
        for test_exe_info in self.test_execution_info_list:
            #Filter None Values from list
            filtered_list = list()
            filtered_list =  [i for i in test_exe_info if i is not None] 

            if "DUT Type" in filtered_list or "VIF Name" in filtered_list or "VIF Spec" in filtered_list  or "Vendor" in filtered_list:
                self.op_worksheet_obj.write(row,1,filtered_list[1] ,self.odd_format)
                self.op_worksheet_obj.merge_range(row, 1,row,10, filtered_list[1], self.odd_format)
                self.op_worksheet_obj.write(row,0,filtered_list[0] ,self.odd_format) 
                

            if "Summary:" in  filtered_list:
                self.op_worksheet_obj.write(row, 0, filtered_list[0],self.odd_format)
                self.op_worksheet_obj.merge_range(row, 1,row,10,filtered_list[1],self.odd_format)         
                #self.op_worksheet_obj.write_url(row,1, filtered_list[1] ,self.odd_format) 
            
            if "Selected Moi and Testcase Status" in  filtered_list:
                self.op_worksheet_obj.write(row, 0, filtered_list[0],self.odd_format)
                self.op_worksheet_obj.merge_range(row, 1,row,10,filtered_list[1],self.odd_format)         
                self.op_worksheet_obj.write(row,1, filtered_list[1] ,self.odd_format) 
           

            if filtered_list ==  self.header_format :
                for each_val in filtered_list:
                    self.op_worksheet_obj.write(row, col, each_val, self.even_format)  
                    self.op_worksheet_obj.set_column(row, col, 12)
                    col += 1 
                col =0  

            else:
                for each_val in filtered_list : 
                    each_val = str(each_val)
                    if 'Test Execution Info' in each_val:
                        self.op_worksheet_obj.merge_range(row, col, row, 10, 'Test Execution Info_'+self.port_name,self.head_format)
                        col = 0
                    else:
                        self.op_worksheet_obj.write(row, col, each_val, self.odd_format)  
                        self.op_worksheet_obj.set_column('A:A', 30)  
                        self.op_worksheet_obj.set_column('B:B', 30) 
                        col += 1 
                col =0            
            row += 1      

    def get_updated_sum_details(self):     
        wk = load_workbook(self.temp_file_path)
        #To get summary details data
        wk_obj =wk["Summary"]   

        sum_sheet_cells = []    
        #iterate each row and convert the each row data's as a tuple format  
        for rows in wk_obj.iter_rows():
            row_cells = []
            for cell in rows:
                row_cells.append(cell.value)
                # create list which will contain all the row data's as tuple format
            sum_sheet_cells.append(list(row_cells))   

        self.up_summary_sheet_data = sum_sheet_cells
        print(self.up_summary_sheet_data)
        for up_val in self.up_summary_sheet_data:
            fil_list =  [i for i in up_val if i is not None] 
            self.fil_sum_data.append(fil_list)
        print(self.fil_sum_data)  

    def over_write_xls(self):
        
        if os.path.exists(self.temp_file_path):
            os.remove(self.temp_file_path)
        self.up_workbook_obj = xlsxwriter.Workbook(self.temp_file_path)
        self.up_worksheet_obj = self.up_workbook_obj.add_worksheet("Summary")  

        self.head_format = self.up_workbook_obj.add_format(
            {'bg_color': '#702D2F', 'border': 1, 'font_color': '#FFFFFF', 'align': 'center',
                 'valign': 'vcenter', 'border_color': '#FFFFFF'})
        self.odd_format = self.up_workbook_obj.add_format(
            {'bg_color': '#F0E6E5', 'border': 1})      
        self.even_format = self.up_workbook_obj.add_format(
            {'bg_color': '#702D2F', 'border': 1,'font_color': '#FFFFFF'}) 

      
        overall_count = len(self.fil_sum_data)
        self.fil_sum_data.append([])
    
        temp_data = list()
        for overwrite_data in self.fil_sum_data:
             
            list_size = len(overwrite_data)
            if list_size == 0 :
                temp_data.append(overwrite_data)
                temp_data.remove([])
                
                if ['Test Setup Information'] == temp_data[0]:
                    self.xls_overwrite_testInformation_writter(temp_data)
                    size = len(temp_data)+1
                    self.tot_count = size
                    temp_data.clear()
                  
                else :
                    self.xls_overwrite_testExeution_info_writter(self.tot_count,temp_data)
                    size = len(temp_data)+1
                    self.tot_count = self.tot_count+size
                    temp_data.clear()               
            else:
                temp_data.append(overwrite_data) 
        self.xls_overwrite_testExeution_info_writter(self.tot_count,self.test_execution_info_list) 
        self.up_workbook_obj.close()       


    def xls_overwrite_testInformation_writter(self,up_set_test_info_order):         
        if os.path.isfile(self.temp_file_path): 
            pass
        else:              
            self.up_worksheet_obj.merge_range(0, 0, 0, 10, "Test Setup Information",self.head_format)
            row = 1
            col = 0   
            up_set_test_info_order.remove(['Test Setup Information'])                             
            for test_info_list in up_set_test_info_order:             
                for each_val in test_info_list:                      
                    if col == 1:           
                        self.up_worksheet_obj.merge_range(row, col,row,col+9, each_val, self.odd_format)   
                    else:
                        self.up_worksheet_obj.write(row, col, each_val, self.odd_format)  
                        self.up_worksheet_obj.set_column('A:A', 30)
                    col += 1 
                row += 1
                col =0
            print(row)   
            self.xls_testExeution_info_writter(row)   
             


    def xls_overwrite_testExeution_info_writter(self,tot_row_count,up_test_execution_info_list):   
        row = tot_row_count+1
        col = 0
        for test_exe_info in up_test_execution_info_list:
            #Filter None Values from list
            filtered_list = list()
            filtered_list =  [i for i in test_exe_info if i is not None] 

            if "DUT Type" in filtered_list or "VIF Name" in filtered_list or "VIF Spec" in filtered_list  or "Vendor" in filtered_list:
                self.up_worksheet_obj.write(row,1,filtered_list[1] ,self.odd_format)
                self.up_worksheet_obj.merge_range(row, 1,row,10, filtered_list[1], self.odd_format)
                self.up_worksheet_obj.write(row,0,filtered_list[0] ,self.odd_format) 
                

            if "Summary:" in  filtered_list:
                self.up_worksheet_obj.write(row, 0, filtered_list[0],self.odd_format)
                self.up_worksheet_obj.merge_range(row, 1,row,10,filtered_list[1],self.odd_format)         
                self.up_worksheet_obj.write_url(row,1, filtered_list[1] ,self.odd_format) 

            if "Selected Moi and Testcase Status" in  filtered_list:
                self.up_worksheet_obj.write(row, 0, filtered_list[0],self.odd_format)
                self.up_worksheet_obj.merge_range(row, 1,row,10,filtered_list[1],self.odd_format)         
                self.up_worksheet_obj.write(row,1, filtered_list[1] ,self.odd_format)     

            if filtered_list ==  self.header_format :
                for each_val in filtered_list:
                    self.up_worksheet_obj.write(row, col, each_val, self.even_format)  
                    self.up_worksheet_obj.set_column(row, col, 12)
                    col += 1 
                col =0  

            else:
                for each_val in filtered_list : 
                    each_val = str(each_val)
                    if 'Test Execution Info' in each_val:
                        if 'Test Execution Info' == each_val:
                            self.up_worksheet_obj.merge_range(row, col, row, 10, 'Test Execution Info_'+self.port_name,self.head_format)
                            col = 0
                        else:
                            self.up_worksheet_obj.merge_range(row, col, row, 10, each_val,self.head_format)
                            col = 0
                    else:
                        self.up_worksheet_obj.write(row, col, each_val, self.odd_format)  
                        self.up_worksheet_obj.set_column('A:A', 30)  
                        self.up_worksheet_obj.set_column('B:B', 30) 
                        col += 1 
                col =0            
            row += 1          




                             
def main():
   
    port_name       = sys.argv[1] 
    reader = CreateXlsSummaryReport(port_name)

if __name__ == "__main__":
    main()
   
   